"""
Pull submission logs from Render and populate extras/onboard_metrics.xlsx (Sheet: intake).

Setup:
    conda run -n ds pip install requests openpyxl
    export RENDER_API_KEY=rnd_...            # Render dashboard → Account Settings → API Keys
    export RENDER_SERVICE_ID=srv-d97lh5etrd3s7399aqc0   # career-transition-intake service ID

Usage:
    conda run -n ds python extras/pull_render_logs.py              # last 30 days
    conda run -n ds python extras/pull_render_logs.py --days 7     # last 7 days
    conda run -n ds python extras/pull_render_logs.py --since 2026-07-01
"""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing dependency — run: conda run -n ds pip install requests openpyxl")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency — run: conda run -n ds pip install requests openpyxl")

XLSX            = Path(__file__).parent / "onboard_metrics.xlsx"
API_BASE        = "https://api.render.com/v1"
INTAKE_SHEET    = "intake"
PROCESSED_SHEET = "processed"

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="F4F6F9")

INTAKE_COLS = [
    "Submission Date",
    "Submission Time (UTC)",
    "Client Name",
    "Client Email",
    "Target Domain",
    "Uploads",
    "Time on Form (min)",
    "Server Processing (sec)",
    "Email Status",
    "Attachment",
]

PROCESSED_COLS = [
    "Client Name",
    "Processing Start",
    "Processing End",
    "Duration (min)",
    "Output File",
    "Notes",
]

ANSI = re.compile(r"\x1b\[[0-9;]*m")


# ── Render API helpers ────────────────────────────────────────────────────────

def _auth_headers():
    key = os.environ.get("RENDER_API_KEY", "").strip()
    if not key:
        sys.exit(
            "RENDER_API_KEY not set.\n"
            "Get one at: Render dashboard → Account Settings → API Keys\n"
            "Then: export RENDER_API_KEY=rnd_..."
        )
    return {"Authorization": f"Bearer {key}", "Accept": "application/json"}


def _get_owner_and_service(headers):
    """Return (owner_id, service_id) for the career-transition-intake service."""
    service_id = os.environ.get("RENDER_SERVICE_ID", "").strip()
    owner_id   = os.environ.get("RENDER_OWNER_ID", "").strip()

    r = requests.get(f"{API_BASE}/services?limit=100", headers=headers, timeout=15)
    r.raise_for_status()
    items = r.json() if isinstance(r.json(), list) else r.json().get("services", [])

    for item in items:
        svc  = item.get("service", item)
        name = svc.get("name", "")
        sid  = svc.get("id", "")
        oid  = svc.get("ownerId", "")

        if not service_id and "career-transition-intake" in name.lower():
            service_id = sid
            print(f"  Found service: {name!r} → {sid}")

        if not owner_id and sid == service_id:
            owner_id = oid

        # Pick up owner from any service if still unknown
        if not owner_id and oid:
            owner_id = oid

    if not service_id:
        sys.exit(
            "Could not find 'career-transition-intake' service.\n"
            "Set RENDER_SERVICE_ID=srv-... (visible in Render dashboard URL) and retry."
        )
    if not owner_id:
        sys.exit("Could not determine owner/workspace ID. Set RENDER_OWNER_ID=tea-... and retry.")

    return owner_id, service_id


def _fetch_logs_page(owner_id, service_id, start, end, headers):
    """Fetch one page of app logs filtered to Submission lines."""
    params = [
        ("ownerId",   owner_id),
        ("resource",  service_id),
        ("startTime", start),
        ("endTime",   end),
        ("type",      "app"),
        ("text",      "Submission"),
        ("limit",     "100"),
        ("direction", "backward"),
    ]
    r = requests.get(f"{API_BASE}/logs", headers=headers, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def _fetch_all_submission_logs(owner_id, service_id, since_dt, until_dt, headers):
    """Paginate through all Submission log lines in the given window."""
    start = since_dt.isoformat()
    end   = until_dt.isoformat()
    entries = []
    pages   = 0

    while pages < 50:
        data = _fetch_logs_page(owner_id, service_id, start, end, headers)
        batch = data.get("logs") or []
        entries.extend(batch)
        pages += 1

        if not data.get("hasMore") or not batch:
            break
        # Render paginates backward: use nextStartTime / nextEndTime for next page
        end   = data.get("nextEndTime",   end)
        start = data.get("nextStartTime", start)

    return entries


# ── Log parsing ───────────────────────────────────────────────────────────────

RE_RECEIVED = re.compile(
    r"Submission received: "
    r"name=(?P<name>.+?) "
    r"email=(?P<email>\S+) "
    r"target=(?P<target>.+?) "
    r"uploads=(?P<uploads>\d+)"
    r"(?:\s+time_on_form=(?P<time_on_form>\d+))?"
)
RE_COMPLETE = re.compile(
    r"Submission complete: "
    r"name=(?P<name>.+?) "
    r"email_status=(?P<email_status>\S+) "
    r"attachment=(?P<attachment>\S+)"
)


def _parse_ts(entry):
    raw = entry.get("timestamp") or entry.get("time") or entry.get("createdAt") or ""
    if raw:
        try:
            return datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        except ValueError:
            pass
    return None


def _parse_entries(entries):
    received_list = []
    complete_list = []

    for e in entries:
        msg = ANSI.sub("", e.get("message") or "").strip()
        dt  = _parse_ts(e)

        m = RE_RECEIVED.search(msg)
        if m:
            d = m.groupdict()
            d["_dt"] = dt
            received_list.append(d)
            continue

        m = RE_COMPLETE.search(msg)
        if m:
            d = m.groupdict()
            d["_dt"] = dt
            complete_list.append(d)

    rows = []
    used = set()

    for recv in received_list:
        name    = recv["name"]
        dt_recv = recv["_dt"]
        proc_secs    = ""
        email_status = ""
        attachment   = ""

        for i, comp in enumerate(complete_list):
            if i in used or comp["name"] != name:
                continue
            dt_comp = comp.get("_dt")
            if dt_recv and dt_comp and abs((dt_comp - dt_recv).total_seconds()) < 600:
                used.add(i)
                proc_secs    = round((dt_comp - dt_recv).total_seconds(), 1)
                email_status = comp.get("email_status", "")
                attachment   = comp.get("attachment", "")
                break

        tof_raw = recv.get("time_on_form")
        tof_min = round(int(tof_raw) / 60, 1) if tof_raw else ""

        rows.append({
            "Submission Date":           dt_recv.strftime("%Y-%m-%d") if dt_recv else "",
            "Submission Time (UTC)":     dt_recv.strftime("%H:%M:%S") if dt_recv else "",
            "Client Name":               name,
            "Client Email":              recv.get("email", ""),
            "Target Domain":             recv.get("target", ""),
            "Uploads":                   int(recv.get("uploads", 0)),
            "Time on Form (min)":        tof_min,
            "Server Processing (sec)":   proc_secs,
            "Email Status":              email_status,
            "Attachment":                attachment,
        })

    return rows


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _load_or_create():
    if XLSX.exists():
        return openpyxl.load_workbook(XLSX)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _ensure_sheet(wb, name, cols):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(1, ci, col)
            c.font      = HEADER_FONT
            c.fill      = HEADER_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes    = "A2"
        ws.row_dimensions[1].height = 30
    return wb[name]


def _existing_keys(ws):
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:
            keys.add((str(row[0]), str(row[2])))
    return keys


def _append_rows(ws, rows, cols, existing_keys):
    added = 0
    for row_data in rows:
        key = (row_data.get("Submission Date", ""), row_data.get("Client Name", ""))
        if key in existing_keys:
            print(f"  Skipping duplicate: {key[1]} on {key[0]}")
            continue
        r = ws.max_row + 1
        for ci, col in enumerate(cols, 1):
            c = ws.cell(r, ci, row_data.get(col, ""))
            if r % 2 == 0:
                c.fill = ALT_FILL
        existing_keys.add(key)
        added += 1
    return added


def _autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 45)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Pull Render submission logs → onboard_metrics.xlsx")
    parser.add_argument("--days",  type=int, default=30,
                        help="Days back to fetch (default: 30)")
    parser.add_argument("--since", help="Fetch from this date, e.g. 2026-07-01 (overrides --days)")
    args = parser.parse_args()

    until = datetime.now(timezone.utc)
    since = (
        datetime.fromisoformat(args.since).replace(tzinfo=timezone.utc)
        if args.since
        else until - timedelta(days=args.days)
    )

    print(f"Fetching Render logs from {since.date()} to {until.date()} …")
    headers            = _auth_headers()
    owner_id, service_id = _get_owner_and_service(headers)
    print(f"  Owner:   {owner_id}")
    print(f"  Service: {service_id}")

    entries = _fetch_all_submission_logs(owner_id, service_id, since, until, headers)
    print(f"  {len(entries)} Submission log line(s) found")

    rows = _parse_entries(entries)
    print(f"  {len(rows)} submission event(s) parsed")

    wb       = _load_or_create()
    ws       = _ensure_sheet(wb, INTAKE_SHEET,    INTAKE_COLS)
    _ensure_sheet(wb, PROCESSED_SHEET, PROCESSED_COLS)

    existing = _existing_keys(ws)
    added    = _append_rows(ws, rows, INTAKE_COLS, existing)

    _autofit(ws)
    _autofit(wb[PROCESSED_SHEET])
    wb.save(XLSX)
    print(f"  {added} new row(s) written → {XLSX}")

    if len(entries) == 0:
        print("\n  Note: no submissions found in this window.")
        print("  The form logs 'Submission received:' on each successful POST.")
        print("  Retry after the first client submits the live form.")


if __name__ == "__main__":
    main()
