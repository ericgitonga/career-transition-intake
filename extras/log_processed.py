"""
Log a plan-generation run to extras/onboard_metrics.xlsx (Sheet: processed).

Run this after generating a client's transition plan PDF.

Usage examples:
    # Provide start time and output path; script computes duration to now:
    python extras/log_processed.py "Jean Kamau" --start "2026-07-14 10:00" --output "Clients/Jean/JK_transition_plan.pdf"

    # Provide start and explicit end:
    python extras/log_processed.py "Jean Kamau" --start "2026-07-14 10:00" --end "2026-07-14 10:43"

    # Provide duration in minutes directly (end defaults to now):
    python extras/log_processed.py "Jean Kamau" --duration 43 --output "Clients/Jean/JK_transition_plan.pdf"

    # With optional notes:
    python extras/log_processed.py "Jean Kamau" --start "2026-07-14 10:00" --notes "Revised Section 2 per client feedback"
"""

import argparse
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency — run: pip install openpyxl")

XLSX  = Path(__file__).parent / "onboard_metrics.xlsx"
SHEET = "processed"

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="F4F6F9")

COLS = [
    "Client Name",
    "Processing Start",
    "Processing End",
    "Duration (min)",
    "Output File",
    "Notes",
]


def _ensure_sheet(wb):
    if SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SHEET)
        for ci, col in enumerate(COLS, 1):
            c = ws.cell(1, ci, col)
            c.font  = HEADER_FONT
            c.fill  = HEADER_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
    return wb[SHEET]


def _autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 45)


def _parse_dt(s):
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    sys.exit(f"Cannot parse datetime: {s!r}  — use YYYY-MM-DD HH:MM")


def main():
    parser = argparse.ArgumentParser(description="Log plan generation time → onboard_metrics.xlsx")
    parser.add_argument("client_name", help="Client full name (must match intake sheet)")
    parser.add_argument("--start",    help="Processing start, e.g. '2026-07-14 10:00'")
    parser.add_argument("--end",      help="Processing end (default: now)")
    parser.add_argument("--duration", type=float, help="Duration in minutes (alternative to --start/--end pair)")
    parser.add_argument("--output",   default="", help="Path to the generated plan PDF")
    parser.add_argument("--notes",    default="", help="Any free-form notes about this run")
    args = parser.parse_args()

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    if args.duration is not None:
        end_str   = args.end or now_str
        dt_end    = _parse_dt(end_str)
        dt_start  = dt_end - timedelta(minutes=args.duration)
        start_str = dt_start.strftime("%Y-%m-%d %H:%M")
        duration  = round(args.duration, 1)
    elif args.start:
        start_str = args.start
        end_str   = args.end or now_str
        dt_start  = _parse_dt(start_str)
        dt_end    = _parse_dt(end_str)
        duration  = round((dt_end - dt_start).total_seconds() / 60, 1)
    else:
        parser.error("Provide --start or --duration.")

    if XLSX.exists():
        wb = openpyxl.load_workbook(XLSX)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = _ensure_sheet(wb)
    r  = ws.max_row + 1
    for ci, val in enumerate(
        [args.client_name, start_str, end_str, duration, args.output, args.notes], 1
    ):
        c = ws.cell(r, ci, val)
        if r % 2 == 0:
            c.fill = ALT_FILL

    _autofit(ws)
    wb.save(XLSX)
    print(
        f"Logged: {args.client_name!r} | {start_str} → {end_str} "
        f"| {duration} min | {XLSX}"
    )


if __name__ == "__main__":
    main()
